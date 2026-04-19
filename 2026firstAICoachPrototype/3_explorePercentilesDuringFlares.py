"""
For the baseline experiment, for each pain type:
- Find days with pain >= 3.2 (after the chronic window, aligned with RF features).
- Merge into episodes: consecutive high-pain days in time order are one episode if each
  gap to the next high-pain day is at most 1 month (31 days); keep only the first day per episode.
- For each such day, compute rank percentiles of RF-selected stressors (_acute/_chronic/_acwr)
  vs the full train+val+test distribution (same slice as 3_runAll_new_trainValTest_SHAP_allTest).
- Saves one Excel per pain type under results/combinedAnalysis/explorePercentilesDuringFlares/:
  rows = stressors (human-readable), columns = two header rows (dates; Flare day 1, 2, ...).
  Percentiles are rounded to 1 decimal.
"""

import json
import re
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from scipy.stats import percentileofscore

from libraries.featureEngineeringFunctions import featureEngineering
from libraries.trainRandomForest import randomForestTrain
from libraries.dataLoader import load_and_preprocess_data
from libraries.run_context import RESULTS_ROOT

BASELINE_PARAMS_DIR = RESULTS_ROOT / "baseline" / "hyperparametersGridSearchResults"
OUTPUT_DIR = RESULTS_ROOT / "combinedAnalysis" / "explorePercentilesDuringFlares"
SILENT_MODE = 1
TEST_ON_TEST_DATASET = True
FILE_PATH = "newDataset.pkl"
STRESSOR_VARS_MINMAX_SCALER = 0
PAIN_REMOVE_OUTLIERS = 0
SPLIT_PERCENT_TRAINVAL_TEST = 0.75
SPLIT_PERCENT_TRAIN_VAL = 0.7
PAIN_TYPES = ["kneePain", "facePain", "armPain"]

PAIN_FLARE_THRESHOLD = 3.2
MAX_GAP_DAYS_WITHIN_EPISODE = 31
PERCENTILE_DECIMALS = 1

# When True, feature windows are fixed at 15/35 to match row labels in _stressor_row_label.
# When False, ACUTE_WINDOW / CHRONIC_WINDOW come from baseline best_params (may differ).
USE_FIXED_15_35_WINDOWS = True

_STRESSOR_SUFFIXES = ("_acute", "_chronic", "_acwr")

# Base variable name (without _acute/_chronic/_acwr) -> display fragment (lowercase phrase).
_VARIABLE_BASE_DISPLAY: Dict[str, str] = {
    "numberOfSteps": "number of steps",
    "manicTimeRealTime": "time on computer",
    "timeSpentDriving": "time spent driving",
    "cyclingCalories": "cycling active calories",
    "phoneTime": "phone time",
    "numberOfComputerClicksAndKeyStrokes": "number of computer mouse and keyboard strokes",
    "numberOfHeartBeatsAbove110_upperBodyActivity": "upper-body activities cumulative bpm > 110",
    "climbingMaxEffortIntensity": "rock climbing maximum route grade",
    "surfCumBpmAbove110": "surfing cumulative bpm > 110",
}


def _split_stressor_suffix(col_name: str) -> Tuple[str, str]:
    """Return (base_var_name, suffix) where suffix is '_acute', '_chronic', or '_acwr'."""
    for s in _STRESSOR_SUFFIXES:
        if col_name.endswith(s):
            return col_name[: -len(s)], s
    return col_name, ""


def _display_base(base: str) -> str:
    if base in _VARIABLE_BASE_DISPLAY:
        return _VARIABLE_BASE_DISPLAY[base]
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", base).lower()
    return s


def _stressor_row_label(col_name: str) -> str:
    """
    Row label rules:
    - _chronic: starts with '35-day chronic'
    - _acute: starts with '15-day chronic' (per spec)
    - _acwr: starts with '15-day/35-day', ends with 'ratio'
    """
    base, suf = _split_stressor_suffix(col_name)
    mid = _display_base(base)
    if suf == "_chronic":
        return f"35-day chronic {mid}"
    if suf == "_acute":
        return f"15-day chronic {mid}"
    if suf == "_acwr":
        return f"15-day/35-day {mid} ratio"
    return mid


def _load_best_params(pain_type: str) -> dict:
    json_path = BASELINE_PARAMS_DIR / f"best_params_{pain_type}.json"
    if not json_path.exists():
        raise FileNotFoundError(f"Best params not found: {json_path}. Run 2_baseline.py first.")
    with json_path.open("r") as f:
        return json.load(f)


def _is_rf_stressor_column(name: str) -> bool:
    return any(name.endswith(s) for s in _STRESSOR_SUFFIXES)


def _first_day_per_episode(sorted_unique_dates: List[pd.Timestamp]) -> List[pd.Timestamp]:
    """
    Dates are high-pain days in ascending order. Episodes merge high-pain days when the
    gap between consecutive high-pain days is <= MAX_GAP_DAYS_WITHIN_EPISODE days.
    Return the first calendar day of each episode.
    """
    if not sorted_unique_dates:
        return []
    out: List[pd.Timestamp] = []
    episode_first = sorted_unique_dates[0]
    prev = sorted_unique_dates[0]
    for d in sorted_unique_dates[1:]:
        gap = (d - prev).days
        if gap <= MAX_GAP_DAYS_WITHIN_EPISODE:
            prev = d
            continue
        out.append(episode_first)
        episode_first = d
        prev = d
    out.append(episode_first)
    return out


def _flare_episode_first_days(full_fe_slice: pd.DataFrame, pain_type: str) -> List[pd.Timestamp]:
    pain = full_fe_slice[pain_type]
    high = pain[pain >= PAIN_FLARE_THRESHOLD].index
    if len(high) == 0:
        return []
    unique_sorted = sorted(pd.to_datetime(high).unique())
    return _first_day_per_episode(list(unique_sorted))


def _percentile_rank_full_distribution(series_all: np.ndarray, value) -> float:
    """Same convention as 3_runAll_new_trainValTest_SHAP_allTest (percentileofscore, kind='rank')."""
    return round(
        float(percentileofscore(np.asarray(series_all), value, kind="rank")),
        PERCENTILE_DECIMALS,
    )


def _flare_column_multiindex(flare_days: List[pd.Timestamp]) -> pd.MultiIndex:
    dates = [
        str(pd.Timestamp(d).date()) if d is not None else ""
        for d in flare_days
    ]
    flare_labels = [f"Flare day {i + 1}" for i in range(len(flare_days))]
    return pd.MultiIndex.from_arrays([dates, flare_labels])


def _full_table_column_multiindex(flare_days: List[pd.Timestamp]) -> pd.MultiIndex:
    """Two header rows: row1 = dates (blank over stressor column), row2 = 'stressor' + Flare day k."""
    dates = [
        str(pd.Timestamp(d).date()) if d is not None else ""
        for d in flare_days
    ]
    flare_labels = [f"Flare day {i + 1}" for i in range(len(flare_days))]
    tuples = [("", "stressor")] + list(zip(dates, flare_labels))
    return pd.MultiIndex.from_tuples(tuples)


def _write_excel_multiindex_columns_no_index(df: pd.DataFrame, path: Path) -> None:
    """
    pandas to_excel does not support MultiIndex columns with index=False (NotImplementedError).
    Write two header rows (MultiIndex levels) then body with openpyxl.
    """
    from openpyxl import Workbook

    cols = df.columns
    if not isinstance(cols, pd.MultiIndex):
        raise TypeError("expected DataFrame with MultiIndex columns")

    wb = Workbook()
    ws = wb.active
    ncols = len(cols)
    nrows = len(df)

    for j in range(ncols):
        col = cols[j]
        t = tuple(col) if isinstance(col, tuple) else (col,)
        c0 = t[0] if len(t) > 0 else ""
        c1 = t[1] if len(t) > 1 else ""
        ws.cell(row=1, column=j + 1, value=None if c0 == "" else c0)
        ws.cell(row=2, column=j + 1, value=None if c1 == "" else c1)

    for i in range(nrows):
        for j in range(ncols):
            val = df.iat[i, j]
            if pd.isna(val):
                val = None
            elif isinstance(val, (np.floating, float)):
                val = float(val)
            elif isinstance(val, (np.integer, int)) and not isinstance(val, bool):
                val = int(val)
            ws.cell(row=3 + i, column=j + 1, value=val)

    wb.save(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df_train_and_val, _df_train, df_val, df_test = load_and_preprocess_data(
        FILE_PATH,
        SPLIT_PERCENT_TRAINVAL_TEST,
        STRESSOR_VARS_MINMAX_SCALER,
        PAIN_REMOVE_OUTLIERS,
        SPLIT_PERCENT_TRAIN_VAL,
    )
    df_full = pd.concat([df_train_and_val, df_test], axis=0)

    with tempfile.TemporaryDirectory() as tmp_rf_dir:
        rf_figures_dir = Path(tmp_rf_dir)
        for pain_type in PAIN_TYPES:
            print(f"--- {pain_type} ---")
            params = _load_best_params(pain_type)
            if USE_FIXED_15_35_WINDOWS:
                acute_window = 15
                chronic_window = 35
            else:
                acute_window = int(params["ACUTE_WINDOW"])
                chronic_window = int(params["CHRONIC_WINDOW"])
            candidates = params["CANDIDATES"]
            only_use_acwr = bool(params["ONLY_USE_ACWR"])
            include_nb_previous_day_pain = int(params.get("INCLUDE_NB_PREVIOUS_DAY_PAIN", 0))

            _, _, top_features = randomForestTrain(
                df_train_and_val,
                pain_type,
                params,
                SILENT_MODE,
                TEST_ON_TEST_DATASET,
                figures_dir=rf_figures_dir,
            )

            df_full_fe, _ = featureEngineering(
                df_full.copy(),
                candidates,
                only_use_acwr,
                acute_window,
                chronic_window,
                include_nb_previous_day_pain,
                pain_type,
            )
            full_fe_slice = df_full_fe.iloc[chronic_window:]

            fe_cols = [c for c in top_features if c in full_fe_slice.columns and _is_rf_stressor_column(c)]
            if not fe_cols:
                print(f"  No RF stressor columns (_acute/_chronic/_acwr) in top_features; skipping.")
                continue

            row_labels = [_stressor_row_label(c) for c in fe_cols]

            flare_days = sorted(_flare_episode_first_days(full_fe_slice, pain_type))
            stressor_col = pd.MultiIndex.from_tuples([("", "stressor")])

            if not flare_days:
                print(f"  No flare days (pain >= {PAIN_FLARE_THRESHOLD} after episode grouping).")
                mat = pd.DataFrame(
                    np.array(row_labels, dtype=object).reshape(-1, 1),
                    columns=stressor_col,
                )
            else:
                column_series = []
                for d in flare_days:
                    column_series.append(
                        [
                            _percentile_rank_full_distribution(
                                full_fe_slice[c].values, full_fe_slice.loc[d, c]
                            )
                            for c in fe_cols
                        ]
                    )
                data_block = np.column_stack(column_series)
                left = pd.DataFrame({("", "stressor"): row_labels})
                right = pd.DataFrame(data_block, columns=_flare_column_multiindex(flare_days))
                mat = pd.concat([left, right], axis=1)

            out_path = OUTPUT_DIR / f"percentiles_during_flares_{pain_type}.xlsx"
            _write_excel_multiindex_columns_no_index(mat, out_path)
            print(f"  Saved {out_path} ({len(flare_days)} flare day columns, {len(fe_cols)} stressors)")

    print("\n--- 3_explorePercentilesDuringFlares finished ---")


if __name__ == "__main__":
    main()
